"""
Generates PWA icons and the Google-Sheets-ready devices template (.xlsx)
for the Device Provisioner. Run once:  python build-assets.py
"""
import os
from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
ICONS = os.path.join(HERE, "icons")
os.makedirs(ICONS, exist_ok=True)

NAVY = (11, 31, 51)        # #0b1f33
ORANGE = (224, 122, 31)    # #e07a1f
TEAL = (70, 194, 191)      # #46c2bf

def rounded(size, radius_ratio=0.22):
    img = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)
    r = int(size * radius_ratio)
    d.rounded_rectangle([0, 0, size - 1, size - 1], radius=r, fill=NAVY)
    return img, d

def draw_chip(d, size, maskable=False):
    # padding bigger for maskable so art stays inside the safe zone
    pad = int(size * (0.30 if maskable else 0.22))
    box = [pad, pad, size - pad, size - pad]
    w = box[2] - box[0]
    pin = max(2, int(w * 0.10))
    plen = max(4, int(w * 0.16))
    # pins (legs) on all four sides
    body = [box[0] + plen, box[1] + plen, box[2] - plen, box[3] - plen]
    n = 3
    for i in range(n):
        # top/bottom
        x = body[0] + (i + 1) * (body[2] - body[0]) / (n + 1) - pin / 2
        d.rectangle([x, box[1], x + pin, body[1]], fill=TEAL)
        d.rectangle([x, body[3], x + pin, box[3]], fill=TEAL)
        # left/right
        y = body[1] + (i + 1) * (body[3] - body[1]) / (n + 1) - pin / 2
        d.rectangle([box[0], y, body[0], y + pin], fill=TEAL)
        d.rectangle([body[2], y, box[2], y + pin], fill=TEAL)
    rr = max(3, int(w * 0.06))
    d.rounded_rectangle(body, radius=rr, fill=ORANGE)
    # notch / dot to read as an IC
    dot = max(3, int(w * 0.07))
    cx, cy = body[0] + (body[2] - body[0]) * 0.30, body[1] + (body[3] - body[1]) * 0.30
    d.ellipse([cx - dot, cy - dot, cx + dot, cy + dot], fill=NAVY)

def make_icon(size, name, maskable=False):
    img, d = rounded(size, radius_ratio=0.0 if maskable else 0.22)
    if maskable:  # full-bleed navy background for maskable
        d.rectangle([0, 0, size, size], fill=NAVY)
    draw_chip(d, size, maskable)
    img.save(os.path.join(ICONS, name))
    print("  icons/" + name)

print("Writing icons:")
make_icon(192, "icon-192.png")
make_icon(512, "icon-512.png")
make_icon(512, "icon-maskable-512.png", maskable=True)
make_icon(180, "apple-touch-icon.png")
make_icon(32, "favicon-32.png")

# ---------------------------------------------------------------------------
# devices-template.xlsx — upload to Google Drive and "Open with Google Sheets"
# ---------------------------------------------------------------------------
print("Writing devices-template.xlsx:")
wb = Workbook()

hdr_fill = PatternFill("solid", fgColor="0B1F33")
hdr_font = Font(color="FFFFFF", bold=True, size=11)
center = Alignment(horizontal="left", vertical="center")

# Sheet 1: Devices  (MUST match the app's import/export columns, in order)
dev = wb.active
dev.title = "Devices"
dev_cols = ["Device ID", "Hardware Version", "Manufacturing Number",
            "Firmware ID", "Comments", "32-bit ID", "Created At"]
dev.append(dev_cols)
for c in range(1, len(dev_cols) + 1):
    cell = dev.cell(row=1, column=c)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center
widths = [14, 18, 22, 14, 30, 14, 22]
for i, wdt in enumerate(widths, start=1):
    dev.column_dimensions[chr(64 + i)].width = wdt
dev.freeze_panes = "A2"

# Sheet 2: HardwareVersions  (the dropdown source list)
hw = wb.create_sheet("HardwareVersions")
hw.append(["Hardware Version"])
hw.cell(row=1, column=1).fill = hdr_fill
hw.cell(row=1, column=1).font = hdr_font
hw.column_dimensions["A"].width = 22
for v in ["Rev A", "Rev B", "Rev C", "1.0", "2.0"]:
    hw.append([v])
hw.freeze_panes = "A2"

# Data-validation dropdown on Devices!B (Hardware Version) sourced from the list.
# Google Sheets honours this list validation on import.
dv = DataValidation(type="list",
                    formula1="=HardwareVersions!$A$2:$A$1000",
                    allow_blank=True, showDropDown=False)
dv.error = "Pick a hardware version from the HardwareVersions sheet."
dv.errorTitle = "Unknown hardware version"
dev.add_data_validation(dv)
dv.add("B2:B1000")

out = os.path.join(HERE, "devices-template.xlsx")
wb.save(out)
print("  devices-template.xlsx")
print("Done.")
